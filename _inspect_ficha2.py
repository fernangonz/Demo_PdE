# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("Fichas/Ficha.xlsx", data_only=False)
ws = wb["PI FALTA DE FRANCOBORDO"]
print("merges:", [str(r) for r in ws.merged_cells.ranges])
imgs = list(getattr(ws, "_images", []) or [])
print("n images", len(imgs))
for i, img in enumerate(imgs):
    anchor = getattr(img, "anchor", None)
    fr = getattr(anchor, "_from", None)
    to = getattr(anchor, "_to", None)
    print("img", i, "from", (fr.col, fr.row) if fr else None, "to", (to.col, to.row) if to else None)
    print(" size", getattr(img, "width", None), getattr(img, "height", None))

# theme colors
try:
    theme = wb.loaded_theme
    print("has theme", bool(theme))
except Exception as e:
    print("theme err", e)

for r in range(1, 20):
    for c in range(1, 4):
        cell = ws.cell(r, c)
        fill = cell.fill
        font = cell.font
        align = cell.alignment
        border = cell.border
        fill_info = None
        if fill and fill.fill_type:
            fg = fill.fgColor
            if fg:
                if fg.type == "rgb":
                    fill_info = ("rgb", fg.rgb, fill.fill_type)
                elif fg.type == "theme":
                    fill_info = ("theme", fg.theme, getattr(fg, "tint", None), fill.fill_type)
                else:
                    fill_info = (fg.type, str(fg.value), fill.fill_type)
        font_info = None
        if font:
            fc = None
            if font.color:
                if font.color.type == "rgb":
                    fc = font.color.rgb
                elif font.color.type == "theme":
                    fc = "theme:%s tint=%s" % (font.color.theme, getattr(font.color, "tint", None))
                else:
                    fc = "%s:%s" % (font.color.type, font.color.value)
            font_info = dict(bold=font.bold, color=fc, size=font.size, name=font.name)
        val = cell.value
        if val is None and not fill_info:
            continue
        print(
            "R%sC%s" % (r, c),
            "val=",
            repr(val)[:90] if val is not None else None,
            "fill=",
            fill_info,
            "font=",
            font_info,
            "align=",
            (align.horizontal, align.vertical, align.wrap_text, align.textRotation),
        )

# Check for drawings / ole / equations in xlsx zip
import zipfile
from pathlib import Path
z = zipfile.ZipFile("Fichas/Ficha.xlsx")
names = z.namelist()
print("--- zip entries ---")
for n in names:
    if any(x in n.lower() for x in ("media", "drawing", "chart", "ole", "equation", "vml", "image")):
        print(n, z.getinfo(n).file_size)
