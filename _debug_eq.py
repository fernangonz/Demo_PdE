# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl
from core.modelos import fichas_excel as fe

wb = openpyxl.load_workbook("Fichas/Ficha.xlsx", data_only=False)
ws = wb["PI FALTA DE FRANCOBORDO"]
print("A5", ws["A5"].value)
print("A6", ws["A6"].value)
print("A12", ws["A12"].value)
wb.close()

textos = fe._extraer_textos_dibujo_hoja("PI FALTA DE FRANCOBORDO")
for t in textos:
    print("TD", t.row, t.col, repr(t.texto))

# debug omml on drawing
import zipfile, re
z = zipfile.ZipFile("Fichas/Ficha.xlsx")
xml = z.read("xl/drawings/drawing1.xml").decode("utf-8")
for m in re.finditer(r"<xdr:(oneCellAnchor)\b.*?</xdr:\1>", xml, flags=re.S):
    chunk = m.group(0)
    print("chunk has Fallback", "<mc:Fallback>" in chunk)
    fb = re.search(r"<mc:Fallback>(.*?)</mc:Fallback>", chunk, flags=re.S)
    if fb:
        texts = re.findall(r"<a:t[^>]*>(.*?)</a:t>", fb.group(1))
        print("n a:t", len(texts))
        for i, x in enumerate(texts):
            print(i, repr(x[:120]))
    print("RESULT", repr(fe._texto_desde_omml(chunk)))
