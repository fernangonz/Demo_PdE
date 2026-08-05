"""Modelos de impacto — capa de compatibilidad.

La implementación vive en ``core/modelos/impacto/pi_agitacion/``.
Este módulo mantiene la API anterior para no romper imports existentes.
"""

from __future__ import annotations

from dataclasses import dataclass, field

import pandas as pd

from core.modelos.impacto.pi_agitacion.calcular import calcular
from core.modelos.impacto.pi_agitacion.schemas import (
    ParametrosEntrada,
    ResultadoPIAgitacion,
    SintesisCambios,
)
from core.modelos.impacto.pi_calado import ResultadoPICalado
from core.modelos.impacto.pi_agitacion.schemas import (
    BASELINE_YEAR,
    MODO_FALLO_DEFAULT,
    VARIABLE_DEFAULT,
)

# Alias históricos
MODO_FALLO_AGITACION = MODO_FALLO_DEFAULT
VARIABLE_OLEAJE = VARIABLE_DEFAULT
ParametrosPIAgitacion = ParametrosEntrada
ParametrosAgitacion = ParametrosEntrada


@dataclass
class IndicadorEstado:
    nombre: str
    seleccionado: bool
    descartado: bool = False


ResumenCambios = SintesisCambios

HORIZONTE_POR_ANIO = {
    2040: "2020-2040",
    2050: "2030-2050",
    2060: "2040-2060",
    2070: "2050-2070",
    2080: "2060-2080",
    2090: "2070-2090",
    2100: "2080-2100",
}


def _parse_etiqueta_escenario(
    escenario: str,
    *,
    baseline_year: int = BASELINE_YEAR,
) -> tuple[str, str, int | None, tuple[int, int]]:
    """Devuelve escenario, horizonte, año representativo y clave de orden."""
    texto = str(escenario).strip()
    if texto.lower() in ("histórico", "historico"):
        return "Histórico", "1995-2014", baseline_year, (0, 0)

    partes = texto.rsplit(" ", 1)
    if len(partes) == 2 and partes[1].isdigit():
        esc = partes[0]
        anio = int(partes[1])
        horizonte = HORIZONTE_POR_ANIO.get(anio, "")
        esc_ord = 0 if "2-4.5" in esc or esc.upper().startswith("SSP2") else 1
        return esc, horizonte, anio, (anio, esc_ord)

    return texto, "", None, (9999, 0)


MODOS_FALLO_PLANTILLA = [
    "PI Exceso de Oleaje",
    "PI FALTA DE FRANCOBORDO",
    "PI FALTA DE CALADO",
    "PI Exceso de Viento",
    "PI Exceso de Corriente",
    "PI Visibilidad reducida",
    "PI Inundación costera",
    "PI Exceso de precipitación",
    "OPEX FALTA DE CALADO",
    "CAPEX FALTA DE CALADO",
]

PLANTILLA_FILAS_RESUMEN = [
    ("Histórico", "1995-2014", 2005),
    ("SSP2-4.5", "2020-2040", 2040),
    ("SSP2-8.5", "2020-2040", 2040),
    ("SSP2-4.5", "2030-2050", 2050),
    ("SSP2-8.5", "2030-2050", 2050),
    ("SSP2-4.5", "2040-2060", 2060),
    ("SSP2-8.5", "2040-2060", 2060),
    ("SSP2-4.5", "2050-2070", 2070),
    ("SSP2-8.5", "2050-2070", 2070),
    ("SSP2-4.5", "2060-2080", 2080),
    ("SSP2-8.5", "2060-2080", 2080),
    ("SSP2-4.5", "2070-2090", 2090),
    ("SSP2-8.5", "2070-2090", 2090),
    ("SSP2-4.5", "2080-2100", 2100),
    ("SSP2-8.5", "2080-2100", 2100),
]

COL_CAMBIO = "Cambio respecto al hist\u00f3rico"
COL_INTERP = "Interpretaci\u00f3n"

# Prefijos de columnas PI precipitaci\u00f3n (umbral mm en par\u00e9ntesis; din\u00e1micos).
PREF_PRECIP_CAMBIO = "Cambio respecto al hist\u00f3rico ("
PREF_PRECIP_INTERP = "Interpretaci\u00f3n ("
# Un solo indicador: mismos nombres que el modo est\u00e1ndar (sin sufijo).
COL_PRECIP_CAMBIO_BARE = COL_CAMBIO
COL_PRECIP_INTERP_BARE = COL_INTERP
SUBCOLS_MODO_ESTANDAR = (COL_CAMBIO, COL_INTERP)
# Fallback est\u00e1tico (1 mm / 20 mm t\u00edpicos de Excel 4 con 2 indicadores).
SUBCOLS_MODO_PRECIP = (
    "Cambio respecto al hist\u00f3rico (1 mm)",
    "Interpretaci\u00f3n (1 mm)",
    "Cambio respecto al hist\u00f3rico (20 mm)",
    "Interpretaci\u00f3n (20 mm)",
)
SUBCOLS_MODO_PRECIP_UNO = (COL_PRECIP_CAMBIO_BARE, COL_PRECIP_INTERP_BARE)


@dataclass
class ResumenIteracion:
    activo: str
    tipo_uo: str
    modo_fallo: str
    variable_climatica: str
    umbral: str
    indicador_seleccionado: str
    percentil: str
    indicadores: list[IndicadorEstado]
    tabla_resultado: pd.DataFrame
    numero: int = 1
    origen_regla: str = "diagrama"
    log: list[str] = field(default_factory=list)
    impactos_asociados: int = 0
    advertencia_negativos: str | None = None
    resumen_cambios: ResumenCambios | None = None
    resultados_por_pasos: object | None = None
    # Diagnóstico auditable (contrato IteracionEjecucion)
    estado: str = "ok"
    motivo: str | None = None
    error_code: str | None = None
    familia: str = ""
    motor_id: str = ""
    nombre_motor: str = ""
    tipo_impacto: str = ""
    pasos: list = field(default_factory=list)


@dataclass
class ResumenActivo:
    """Tabla consolidada del activo (plantilla Excel escenarios × modos de fallo)."""

    activo: str
    filas: list[dict[str, object]] = field(default_factory=list)
    modos_fallo: list[str] = field(default_factory=lambda: list(MODOS_FALLO_PLANTILLA))
    # Subcolumnas por modo (2 estándar; 2 o 4 para precipitación según N indicadores).
    subcols_por_modo: dict[str, tuple[str, ...]] = field(default_factory=dict)

    def subcols_de(self, modo: str) -> tuple[str, ...]:
        return self.subcols_por_modo.get(modo) or SUBCOLS_MODO_ESTANDAR

    @property
    def tabla_resumen(self) -> pd.DataFrame:
        return tabla_resumen_activo_a_dataframe(self)

    @property
    def columnas(self) -> list[str]:
        return columnas_export_resumen_activo(
            self.modos_fallo,
            subcols_por_modo=self.subcols_por_modo,
        )


@dataclass
class ResultadosImpacto:
    resumen_iteracion: ResumenIteracion | None = None
    resumenes_iteracion: list[ResumenIteracion] = field(default_factory=list)
    resumen_activo: ResumenActivo | None = None
    comentario: str = ""
    modelo: str = "PI_AGITACION"
    variable: str = VARIABLE_OLEAJE
    referencia: str = f"Histórico ({BASELINE_YEAR})"
    incremento: pd.DataFrame = field(default_factory=pd.DataFrame)
    acumulado: pd.DataFrame = field(default_factory=pd.DataFrame)
    equivalente_anual: pd.DataFrame = field(default_factory=pd.DataFrame)
    detalle_modelo: pd.DataFrame = field(default_factory=pd.DataFrame)
    resultado: ResultadoPIAgitacion | None = None


def _escenario_plantilla_a_datos(escenario: str) -> str:
    if escenario == "SSP2-8.5":
        return "SSP5-8.5"
    return escenario


def _norm_col_ascii(nombre: str) -> str:
    return (
        str(nombre)
        .lower()
        .replace("\u00e1", "a")
        .replace("\u00e9", "e")
        .replace("\u00ed", "i")
        .replace("\u00f3", "o")
        .replace("\u00fa", "u")
    )


def _es_col_cambio_precip_paren(nombre: str) -> bool:
    s = str(nombre)
    return s.startswith(PREF_PRECIP_CAMBIO) or s.startswith("Cambio respecto al historico (")


def _es_col_interp_precip_paren(nombre: str) -> bool:
    s = str(nombre)
    return s.startswith(PREF_PRECIP_INTERP) or s.startswith("Interpretacion (")


def _es_col_cambio_bare(nombre: str) -> bool:
    return _norm_col_ascii(nombre) == _norm_col_ascii(COL_PRECIP_CAMBIO_BARE)


def _es_col_interp_bare(nombre: str) -> bool:
    return _norm_col_ascii(nombre) == _norm_col_ascii(COL_PRECIP_INTERP_BARE)


def _es_tabla_calado(tabla: pd.DataFrame) -> bool:
    return "h" in tabla.columns and "NM" in tabla.columns


def _es_tabla_precipitacion(tabla: pd.DataFrame) -> bool:
    """True si hay >=1 columna Cambio con sufijo (mm) o 1 par bare sin Indicador."""
    n_cambio = sum(1 for c in tabla.columns if _es_col_cambio_precip_paren(c))
    if n_cambio >= 1:
        return True
    # Un indicador: columnas sin sufijo; no confundir con PI estándar (tiene Indicador).
    cols_norm = {_norm_col_ascii(c) for c in tabla.columns}
    if "indicador" in cols_norm:
        return False
    return (
        _norm_col_ascii(COL_PRECIP_CAMBIO_BARE) in cols_norm
        and _norm_col_ascii(COL_PRECIP_INTERP_BARE) in cols_norm
    )


def _subcols_precip_desde_tabla(tabla: pd.DataFrame) -> tuple[str, ...]:
    """Orden del DF: pares Cambio/Interpretacion por indicador (1 o 2)."""
    out: list[str] = []
    for c in tabla.columns:
        s = str(c)
        if (
            _es_col_cambio_precip_paren(s)
            or _es_col_interp_precip_paren(s)
            or _es_col_cambio_bare(s)
            or _es_col_interp_bare(s)
        ):
            out.append(s)
    if len(out) >= 2:
        return tuple(out)
    if _es_tabla_precipitacion(tabla):
        return SUBCOLS_MODO_PRECIP_UNO
    return SUBCOLS_MODO_PRECIP


def _col_por_nombre(tabla: pd.DataFrame, nombre: str) -> str | None:
    if nombre in tabla.columns:
        return nombre
    objetivo = _norm_col_ascii(nombre)
    for c in tabla.columns:
        if _norm_col_ascii(c) == objetivo:
            return str(c)
    return None


def _es_col_cambio_precip(nombre: str) -> bool:
    return _es_col_cambio_precip_paren(nombre) or _es_col_cambio_bare(nombre)


def _es_col_interp_precip(nombre: str) -> bool:
    return _es_col_interp_precip_paren(nombre) or _es_col_interp_bare(nombre)


def _indexar_variaciones_por_modo(
    iteraciones: list[ResumenIteracion],
    *,
    baseline_year: int = BASELINE_YEAR,
) -> dict[tuple[str, str, int], dict[str, object]]:
    """Clave: (modo_fallo, escenario_datos, año) → celdas del modo."""
    indice: dict[tuple[str, str, int], dict[str, object]] = {}
    for r in iteraciones:
        if r.tabla_resultado is None or r.tabla_resultado.empty:
            continue
        if _es_tabla_calado(r.tabla_resultado):
            for _, row in r.tabla_resultado.iterrows():
                esc, _horizonte, anio, _orden = _parse_etiqueta_escenario(
                    str(row.get("Escenario", "")),
                    baseline_year=baseline_year,
                )
                if esc == "Histórico" or anio is None:
                    continue
                h_val = row.get("h")
                indice[(r.modo_fallo, esc, anio)] = {
                    COL_CAMBIO: h_val,
                    COL_INTERP: row.get(COL_INTERP, ""),
                }
            continue
        if _es_tabla_precipitacion(r.tabla_resultado):
            subcols = _subcols_precip_desde_tabla(r.tabla_resultado)
            for _, row in r.tabla_resultado.iterrows():
                esc, _horizonte, anio, _orden = _parse_etiqueta_escenario(
                    str(row.get("Escenario", "")),
                    baseline_year=baseline_year,
                )
                if esc == "Histórico" or anio is None:
                    continue
                # Pares independientes por indicador (nunca sumar).
                indice[(r.modo_fallo, esc, anio)] = {
                    col: row.get(col) for col in subcols
                }
            continue
        for _, row in r.tabla_resultado.iterrows():
            esc, _horizonte, anio, _orden = _parse_etiqueta_escenario(
                str(row.get("Escenario", "")),
                baseline_year=baseline_year,
            )
            if esc == "Histórico" or anio is None:
                continue
            indice[(r.modo_fallo, esc, anio)] = {
                COL_CAMBIO: row.get(COL_CAMBIO),
                COL_INTERP: row.get(COL_INTERP, ""),
            }
    return indice


def _valor_celda_resumen(
    *,
    es_historico: bool,
    cambio: object,
    interpretacion: object,
    es_calado: bool = False,
) -> tuple[str, str]:
    if es_historico:
        return "", ""
    if cambio is None or (isinstance(cambio, float) and pd.isna(cambio)):
        cambio_txt = ""
    elif es_calado and isinstance(cambio, (int, float)):
        cambio_txt = f"{float(cambio):.3f}"
    else:
        cambio_txt = str(int(cambio))
    interp_txt = "" if interpretacion is None else str(interpretacion)
    if interp_txt.lower() in ("referencia",):
        interp_txt = ""
    return cambio_txt, interp_txt


def columnas_export_resumen_activo(
    modos: list[str] | None = None,
    *,
    subcols_por_modo: dict[str, tuple[str, ...]] | None = None,
) -> list[str]:
    modos = modos or MODOS_FALLO_PLANTILLA
    subcols_por_modo = subcols_por_modo or {}
    cols = ["Escenarios", "Horizontes temporales", "Año representativo"]
    for modo in modos:
        subcols = subcols_por_modo.get(modo) or SUBCOLS_MODO_ESTANDAR
        for sub in subcols:
            cols.append(f"{modo} | {sub}")
    return cols


def tabla_resumen_activo_a_dataframe(resumen: ResumenActivo) -> pd.DataFrame:
    columnas = columnas_export_resumen_activo(
        resumen.modos_fallo,
        subcols_por_modo=resumen.subcols_por_modo,
    )
    filas_export: list[dict[str, object]] = []
    for fila in resumen.filas:
        registro = {
            "Escenarios": fila["Escenarios"],
            "Horizontes temporales": fila["Horizontes temporales"],
            "Año representativo": fila["Año representativo"],
        }
        modos = fila.get("modos", {})
        for modo in resumen.modos_fallo:
            datos = modos.get(modo, {})
            for sub in resumen.subcols_de(modo):
                registro[f"{modo} | {sub}"] = datos.get(sub, "")
        filas_export.append(registro)
    return pd.DataFrame(filas_export, columns=columnas)


def construir_tabla_resumen_activo(
    iteraciones: list[ResumenIteracion],
    *,
    baseline_year: int = BASELINE_YEAR,
) -> ResumenActivo | None:
    """Construye la plantilla fija del resumen del activo."""
    # Solo modos calculados OK entran en la plantilla consolidada.
    iteraciones = [
        it for it in iteraciones
        if getattr(it, "estado", "ok") == "ok"
        and it.tabla_resultado is not None
        and not it.tabla_resultado.empty
    ]
    if not iteraciones:
        return None

    indice = _indexar_variaciones_por_modo(iteraciones, baseline_year=baseline_year)
    modos_presentes = {r.modo_fallo for r in iteraciones}
    modos = [m for m in MODOS_FALLO_PLANTILLA if m in modos_presentes]
    extra = sorted(m for m in modos_presentes if m not in MODOS_FALLO_PLANTILLA)
    modos.extend(extra)
    if not modos:
        modos = list(MODOS_FALLO_PLANTILLA)

    # Subcolumnas: precipitacion usa 2 o 4 (Cambio/Interpretacion x 1 o 2 indicadores).
    subcols_por_modo: dict[str, tuple[str, ...]] = {}
    for it in iteraciones:
        if _es_tabla_precipitacion(it.tabla_resultado):
            subcols_por_modo[it.modo_fallo] = _subcols_precip_desde_tabla(it.tabla_resultado)
        else:
            subcols_por_modo.setdefault(it.modo_fallo, SUBCOLS_MODO_ESTANDAR)

    filas: list[dict[str, object]] = []
    for esc, horizonte, anio in PLANTILLA_FILAS_RESUMEN:
        es_historico = esc == "Histórico"
        esc_datos = _escenario_plantilla_a_datos(esc)
        modos_valores: dict[str, dict[str, str]] = {}
        for modo in modos:
            subcols = subcols_por_modo.get(modo) or SUBCOLS_MODO_ESTANDAR
            if es_historico:
                modos_valores[modo] = {sub: "" for sub in subcols}
                continue
            datos = indice.get((modo, esc_datos, anio), {})
            es_calado = "Calado" in modo
            celdas: dict[str, str] = {}
            if any(_es_col_cambio_precip(s) for s in subcols):
                for sub in subcols:
                    val = datos.get(sub, "")
                    if _es_col_cambio_precip(sub) and val != "" and val is not None:
                        try:
                            celdas[sub] = str(int(val))
                        except (TypeError, ValueError):
                            celdas[sub] = str(val)
                    else:
                        txt = "" if val is None else str(val)
                        if txt.lower() in ("referencia",):
                            txt = ""
                        celdas[sub] = txt
            else:
                cambio_txt, interp_txt = _valor_celda_resumen(
                    es_historico=False,
                    cambio=datos.get(COL_CAMBIO),
                    interpretacion=datos.get(COL_INTERP),
                    es_calado=es_calado,
                )
                celdas = {COL_CAMBIO: cambio_txt, COL_INTERP: interp_txt}
            modos_valores[modo] = celdas

        filas.append({
            "Escenarios": esc,
            "Horizontes temporales": horizonte,
            "Año representativo": anio,
            "modos": modos_valores,
        })

    return ResumenActivo(
        activo=iteraciones[0].activo,
        filas=filas,
        modos_fallo=modos,
        subcols_por_modo=subcols_por_modo,
    )


def html_tabla_resumen_activo(resumen: ResumenActivo) -> str:
    """HTML con cabeceras agrupadas como la plantilla Excel."""
    import html as html_lib

    modos = resumen.modos_fallo

    def esc(val: object) -> str:
        if val is None:
            return ""
        return html_lib.escape(str(val))

    total_subcols = sum(len(resumen.subcols_de(m)) for m in modos)

    filas_html: list[str] = []
    for fila in resumen.filas:
        celdas = [
            f"<td>{esc(fila['Escenarios'])}</td>",
            f"<td>{esc(fila['Horizontes temporales'])}</td>",
            f"<td>{esc(fila['Año representativo'])}</td>",
        ]
        modos_datos = fila.get("modos", {})
        for modo in modos:
            datos = modos_datos.get(modo, {})
            for sub in resumen.subcols_de(modo):
                celdas.append(f"<td>{esc(datos.get(sub, ''))}</td>")
        filas_html.append(f"<tr>{''.join(celdas)}</tr>")

    subcols = "".join(
        "".join(f"<th>{esc(sub)}</th>" for sub in resumen.subcols_de(modo))
        for modo in modos
    )
    modos_hdr = "".join(
        f'<th colspan="{len(resumen.subcols_de(modo))}">{esc(modo)}</th>'
        for modo in modos
    )

    return f"""
<div class="resumen-activo-tabla">
<style>
.resumen-activo-tabla table {{
  border-collapse: collapse;
  width: 100%;
  font-size: 0.82rem;
}}
.resumen-activo-tabla th,
.resumen-activo-tabla td {{
  border: 1px solid #000;
  padding: 5px 8px;
  text-align: center;
  vertical-align: middle;
}}
.resumen-activo-tabla thead th {{
  background: #ffffff;
  color: #000000;
  font-weight: 600;
}}
</style>
<table>
  <thead>
    <tr>
      <th rowspan="3">Escenarios</th>
      <th rowspan="3">Horizontes temporales</th>
      <th rowspan="3">Año representativo</th>
      <th colspan="{total_subcols}">Modos de fallo / Modos de parada</th>
    </tr>
    <tr>
      {modos_hdr}
    </tr>
    <tr>
      {subcols}
    </tr>
  </thead>
  <tbody>
    {''.join(filas_html)}
  </tbody>
</table>
</div>
""".strip()


def export_resumen_activo_xlsx(resumen: ResumenActivo, *, hoja: str = "Resumen activo") -> bytes:
    """Excel con cabeceras combinadas como la plantilla."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side

    wb = Workbook()
    ws = wb.active
    ws.title = hoja[:31]

    thin = Side(style="thin", color="000000")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    negrita = Font(bold=True)

    modos = resumen.modos_fallo
    total_subcols = sum(len(resumen.subcols_de(m)) for m in modos)

    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    ws.cell(1, 1, "Escenarios")
    ws.merge_cells(start_row=1, start_column=2, end_row=3, end_column=2)
    ws.cell(1, 2, "Horizontes temporales")
    ws.merge_cells(start_row=1, start_column=3, end_row=3, end_column=3)
    ws.cell(1, 3, "Año representativo")
    if total_subcols > 0:
        ws.merge_cells(
            start_row=1,
            start_column=4,
            end_row=1,
            end_column=3 + total_subcols,
        )
        ws.cell(1, 4, "Modos de fallo / Modos de parada")

    col = 4
    for modo in modos:
        subcols = resumen.subcols_de(modo)
        n_sub = len(subcols)
        if n_sub > 1:
            ws.merge_cells(
                start_row=2,
                start_column=col,
                end_row=2,
                end_column=col + n_sub - 1,
            )
        ws.cell(2, col, modo)
        for i, sub in enumerate(subcols):
            ws.cell(3, col + i, sub)
        col += n_sub

    fila_excel = 4
    for fila in resumen.filas:
        ws.cell(fila_excel, 1, fila["Escenarios"])
        ws.cell(fila_excel, 2, fila["Horizontes temporales"])
        ws.cell(fila_excel, 3, fila["Año representativo"])
        col = 4
        modos_datos = fila.get("modos", {})
        for modo in modos:
            datos = modos_datos.get(modo, {})
            for sub in resumen.subcols_de(modo):
                ws.cell(fila_excel, col, datos.get(sub, ""))
                col += 1
        fila_excel += 1

    max_col = max(3, 3 + total_subcols)
    for row in ws.iter_rows(min_row=1, max_row=fila_excel - 1, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = borde
            cell.alignment = centro
            if cell.row <= 3:
                cell.font = negrita

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def resumen_activo_desde_iteraciones(
    iteraciones: list[ResumenIteracion],
) -> ResumenActivo | None:
    return construir_tabla_resumen_activo(iteraciones)


def _motor_meta_desde_resultado(resultado) -> tuple[str, str, str, str]:
    """(motor_id, familia, tipo_impacto, nombre_motor) desde metadatos / ficha."""
    from core.modelos.fichas_modelo import ficha_por_motor, nombre_motor_display

    meta_ej = getattr(resultado, "metadatos_ejecucion", None) or {}
    meta = getattr(resultado, "metadatos", None)
    motor_id = str(
        meta_ej.get("modelo_id", "")
        or getattr(meta, "id", "")
        or ""
    ).strip()
    ficha = ficha_por_motor(motor_id)
    familia = (ficha.familia if ficha else "") or "PI"
    tipo_impacto = (ficha.tipo_impacto if ficha else "") or "ELO"
    nombre = nombre_motor_display(motor_id, familia=familia, tipo_impacto=tipo_impacto)
    return motor_id, familia, tipo_impacto, nombre


def _iteracion_a_resumen(
    resultado: ResultadoPIAgitacion,
    it: object,
    *,
    activo: str,
    tipo_uo: str,
    impactos_asociados: int,
    resultados_por_pasos: object | None,
) -> ResumenIteracion:
    indicadores = [
        IndicadorEstado(i.nombre, i.seleccionado, i.descartado)
        for i in it.indicadores_evaluados
    ]
    advertencia = it.advertencias[0] if it.advertencias else None
    motor_id, familia, tipo_impacto, nombre_motor = _motor_meta_desde_resultado(resultado)
    return ResumenIteracion(
        numero=it.numero,
        activo=activo,
        tipo_uo=tipo_uo,
        modo_fallo=it.modo_fallo,
        variable_climatica=it.variable_climatica,
        umbral=it.umbral,
        indicador_seleccionado=it.indicador_seleccionado,
        percentil=it.percentil,
        origen_regla=it.origen_regla,
        indicadores=indicadores,
        tabla_resultado=it.tabla_resultado,
        impactos_asociados=impactos_asociados,
        advertencia_negativos=advertencia,
        resumen_cambios=it.sintesis_cambios,
        resultados_por_pasos=resultados_por_pasos,
        estado=getattr(it, "estado", "ok") or "ok",
        motivo=getattr(it, "motivo", None),
        error_code=getattr(it, "error_code", None),
        familia=familia,
        motor_id=motor_id,
        nombre_motor=nombre_motor,
        tipo_impacto=tipo_impacto,
    )


def _adaptar_resultado(resultado: ResultadoPIAgitacion) -> ResultadosImpacto:
    if not resultado.iteraciones:
        return ResultadosImpacto(
            comentario=resultado.error or ("Error desconocido." if not resultado.ok else ""),
            resultado=resultado,
        )

    meta = resultado.metadatos_ejecucion
    activo = str(meta.get("activo", ""))
    tipo_uo = str(meta.get("tipo_uo", ""))
    impactos = int(meta.get("impactos_asociados", 0))
    pasos = resultado.resultados_por_pasos

    resumenes = [
        _iteracion_a_resumen(
            resultado,
            it,
            activo=activo,
            tipo_uo=tipo_uo,
            impactos_asociados=impactos,
            resultados_por_pasos=pasos,
        )
        for it in resultado.iteraciones
    ]
    return ResultadosImpacto(
        resumen_iteracion=resumenes[0] if resumenes else None,
        resumenes_iteracion=resumenes,
        resumen_activo=resumen_activo_desde_iteraciones(resumenes),
        comentario=resultado.error or "",
        resultado=resultado,
    )


def resumen_para_ui(resultado: ResultadoPIAgitacion) -> ResumenIteracion | None:
    """Adapta la salida del modelo al formato de la UI Streamlit (primera iteración)."""
    return _adaptar_resultado(resultado).resumen_iteracion


def iteraciones_para_ui(resultado: ResultadoPIAgitacion) -> list[ResumenIteracion]:
    """Todas las iteraciones IM del cálculo."""
    return _adaptar_resultado(resultado).resumenes_iteracion


def _iteracion_calado_a_resumen(
    resultado: ResultadoPICalado,
    it: object,
    *,
    activo: str,
    tipo_uo: str,
    impactos_asociados: int,
) -> ResumenIteracion:
    from core.modelos.fichas_modelo import nombre_motor_display

    indicadores = [
        IndicadorEstado(i.nombre, i.seleccionado, i.descartado)
        for i in getattr(it, "indicadores_evaluados", []) or []
    ]
    motor_id = getattr(it, "motor_id", "") or str(
        (resultado.metadatos_ejecucion or {}).get("modelo_id", "")
    )
    familia = getattr(it, "familia", "") or ""
    tipo_impacto = getattr(it, "tipo_impacto", "") or str(
        (resultado.metadatos_ejecucion or {}).get("tipo_impacto", "")
    )
    nombre_motor = nombre_motor_display(
        motor_id,
        familia=familia,
        tipo_impacto=tipo_impacto,
        modo_fallo=getattr(it, "modo_fallo", "") or "",
    )
    return ResumenIteracion(
        numero=it.numero,
        activo=activo,
        tipo_uo=tipo_uo,
        modo_fallo=it.modo_fallo,
        variable_climatica=getattr(it, "variable_climatica", getattr(it, "variable", "")),
        umbral=getattr(it, "umbral", "") or "",
        indicador_seleccionado=getattr(it, "indicador_seleccionado", "") or "",
        percentil=getattr(it, "percentil", "") or "",
        origen_regla=getattr(it, "origen_regla", "") or "",
        indicadores=indicadores,
        tabla_resultado=it.tabla_resultado,
        impactos_asociados=impactos_asociados,
        advertencia_negativos=None,
        resumen_cambios=None,
        resultados_por_pasos=None,
        estado=getattr(it, "estado", "ok") or "ok",
        motivo=getattr(it, "motivo", None),
        error_code=getattr(it, "error_code", None),
        familia=familia,
        motor_id=motor_id,
        nombre_motor=nombre_motor,
        tipo_impacto=tipo_impacto,
        pasos=list(getattr(it, "pasos", []) or []),
    )


def _ejecuciones_calado_a_resumenes(
    resultado: ResultadoPICalado,
    *,
    activo: str,
    tipo_uo: str,
    impactos_asociados: int,
) -> list[ResumenIteracion]:
    """Prefiere ejecuciones auditables (incluye errores por modo).

    Las omitidas como impacto no factible no se listan como IM calculadas
    (mismo criterio que agitación/francobordo: continue sin iteración).
    """
    from core.modelos.impacto.impactos_no_factibles import CODIGO_NO_FACTIBLE

    ejecuciones = getattr(resultado, "ejecuciones", None) or []
    if ejecuciones:
        return [
            _iteracion_calado_a_resumen(
                resultado,
                ej,
                activo=activo,
                tipo_uo=tipo_uo,
                impactos_asociados=impactos_asociados,
            )
            for ej in ejecuciones
            if not (
                getattr(ej, "estado", "") == "skipped"
                and getattr(ej, "error_code", None) == CODIGO_NO_FACTIBLE
            )
        ]
    return [
        _iteracion_calado_a_resumen(
            resultado,
            it,
            activo=activo,
            tipo_uo=tipo_uo,
            impactos_asociados=impactos_asociados,
        )
        for it in resultado.iteraciones
    ]


def iteraciones_desde_calculo_activo(resultado) -> list[ResumenIteracion]:
    """Todas las iteraciones IM del cálculo unificado por activo."""
    resumenes: list[ResumenIteracion] = []
    ag = getattr(resultado, "resultado_agitacion", None)
    if ag is not None and (ag.ok or ag.iteraciones):
        resumenes.extend(iteraciones_para_ui(ag))

    fb = getattr(resultado, "resultado_francobordo", None)
    if fb is not None and (fb.ok or getattr(fb, "iteraciones", None)):
        resumenes.extend(iteraciones_para_ui(fb))

    prec = getattr(resultado, "resultado_precipitacion", None)
    if prec is not None and (prec.ok or getattr(prec, "iteraciones", None)):
        resumenes.extend(iteraciones_para_ui(prec))

    calado_resultados = []
    for attr in ("resultado_calado_pi", "resultado_calado_opex", "resultado_calado", "resultado_calado_capex"):
        res = getattr(resultado, attr, None)
        if res is None:
            continue
        if res in calado_resultados:
            continue
        # Incluir también resultados con ejecuciones fallidas (diagnóstico).
        if res.ok or getattr(res, "ejecuciones", None):
            calado_resultados.append(res)

    offset = len(resumenes)
    for resultado_cal in calado_resultados:
        meta = resultado_cal.metadatos_ejecucion or {}
        activo = str(meta.get("activo", ""))
        tipo_uo = str(meta.get("tipo_uo", ""))
        impactos = int(meta.get("impactos_asociados", 0))
        bloque = _ejecuciones_calado_a_resumenes(
            resultado_cal,
            activo=activo,
            tipo_uo=tipo_uo,
            impactos_asociados=impactos,
        )
        for idx, resumen in enumerate(bloque, start=1):
            resumen.numero = offset + idx
            resumenes.append(resumen)
        offset = len(resumenes)
    return resumenes


def resumen_activo_desde_calculo_activo(resultado) -> ResumenActivo | None:
    return construir_tabla_resumen_activo(iteraciones_desde_calculo_activo(resultado))


def resumen_activo_para_ui(resultado: ResultadoPIAgitacion) -> ResumenActivo | None:
    """Resumen consolidado del activo (todos los modos de fallo)."""
    return _adaptar_resultado(resultado).resumen_activo


def pi_agitacion(
    info_clima: dict,
    config_puerto: pd.DataFrame | None = None,
    *,
    params: ParametrosEntrada | None = None,
    df_relacion: pd.DataFrame | None = None,
    por_hoja_umbrales: dict[str, pd.DataFrame] | None = None,
) -> ResultadosImpacto:
    """API legacy — delega en ``core.modelos.impacto.pi_agitacion``."""
    p = params or ParametrosEntrada()
    resultado = calcular(
        datos=info_clima,
        params=p,
        info_clima=info_clima,
        config_puerto=config_puerto,
        df_relacion=df_relacion,
        por_hoja_umbrales=por_hoja_umbrales,
    )
    return _adaptar_resultado(resultado)


def nombres_tipos_uo() -> list[str]:
    from core.data_loader import cargar_tipo_de_uo, _normalizar

    fallback = [
        "Graneles Líquidos",
        "Graneles Sólidos",
        "Carga contenerizada",
        "Ro-ro y resto de mercancía general no contenerizada",
        "Pasajeros",
        "Pesquero",
        "Náutico-deportivo",
        "Puerto-ciudad",
    ]
    try:
        df, _ = cargar_tipo_de_uo()
        col = "Tipo de UO"
        if col not in df.columns:
            col = next((c for c in df.columns if "tipo" in _normalizar(c)), None)
        if col:
            vals = [str(v).strip() for v in df[col].dropna() if str(v).strip()]
            if vals:
                return vals
    except (FileNotFoundError, ValueError, OSError):
        pass
    return fallback


def tabla_tipos_terminal() -> pd.DataFrame:
    tipos = nombres_tipos_uo()
    return pd.DataFrame([
        {"Nº": i, "Tipo de UO": t}
        for i, t in enumerate(tipos, start=1)
    ])
