# -*- coding: utf-8 -*-

"""Fichas documentales desde Fichas/Ficha.xlsx (una hoja = un modelo).



Cada hoja se empareja con una entrada del catalogo por nombre normalizado

(motor_nombre, aliases de flujo, modo de fallo). El HTML copia fills, fuentes,

alineacion y merges Excel; las imagenes de esquema quedan aparte (detalle);

ecuaciones en dibujos de texto se incrustan en la tabla.

"""



from __future__ import annotations



import base64

import html as html_lib

import re

import unicodedata

import zipfile

from dataclasses import dataclass

from functools import lru_cache

from pathlib import Path



import pandas as pd



RAIZ_PROYECTO = Path(__file__).resolve().parents[2]

CARPETA_FICHAS = RAIZ_PROYECTO / "Fichas"

ARCHIVO_FICHAS = CARPETA_FICHAS / "Ficha.xlsx"

CARPETA_MEDIA = CARPETA_FICHAS / "_media"



# Notas editoriales / TODO del Excel: no mostrar en la UI.

_PREFIJOS_NOTA_EDITORIAL = (

    "revisar y corregir",

    "especificar:",

)



# OOXML themeClr indices (scheme) -> RGB aproximado Office por defecto.

_THEME_RGB = {

    0: "FFFFFF",  # lt1

    1: "000000",  # dk1

    2: "E7E6E6",  # lt2

    3: "44546A",  # dk2

    4: "5B9BD5",  # accent1

    5: "ED7D31",

    6: "A5A5A5",

    7: "FFC000",

    8: "4472C4",

    9: "70AD47",

}





@dataclass(frozen=True)

class FichaExcelModelo:

    """Contenido de una hoja de Ficha.xlsx asociada a un modelo."""



    hoja: str

    tabla: pd.DataFrame

    html: str

    imagenes: tuple[Path, ...]

    catalogo_id: str | None = None





@dataclass(frozen=True)

class _ImgAnclada:

    """Imagen flotante con ancla (fila/col 1-based Excel)."""



    row: int

    col: int

    path: Path

    data_uri: str





@dataclass(frozen=True)

class _TextoDibujo:

    """Texto de forma/ecuacion anclado a una celda (1-based)."""



    row: int

    col: int

    texto: str





def _normalizar(texto: str) -> str:

    t = unicodedata.normalize("NFKD", str(texto or ""))

    t = t.encode("ascii", "ignore").decode("ascii")

    return re.sub(r"[^a-z0-9]+", "", t.lower())





def _slug(texto: str) -> str:

    t = unicodedata.normalize("NFKD", str(texto or ""))

    t = t.encode("ascii", "ignore").decode("ascii")

    t = re.sub(r"[^A-Za-z0-9]+", "_", t).strip("_")

    return t or "hoja"





def _es_nota_editorial(texto: object) -> bool:

    """True si el texto es una nota de revision/TODO del Excel."""

    t = re.sub(r"\s+", " ", str(texto or "")).strip().lower()

    if not t:

        return False

    return any(t.startswith(p) for p in _PREFIJOS_NOTA_EDITORIAL)





def _aliases_catalogo(entrada) -> set[str]:

    """Nombres normalizados con los que una hoja puede emparejarse."""

    from core.modelos.catalogo_impactos import titulo_modo_impacto

    from core.modelos.flujos import _ALIASES_FLUJO, nombre_esperado_diagrama



    aliases: set[str] = set()

    for bruto in (

        entrada.id,

        entrada.motor_id,

        entrada.motor_nombre,

        entrada.modo_fallo,

        titulo_modo_impacto(entrada),

        f"{entrada.familia} {entrada.modo_fallo}",

        f"{entrada.tipo_impacto} {entrada.modo_fallo}",

        nombre_esperado_diagrama(entrada.diagrama_modelo_id or entrada.motor_id),

    ):

        n = _normalizar(bruto)

        if n:

            aliases.add(n)

    mid = entrada.diagrama_modelo_id or entrada.motor_id

    for a in _ALIASES_FLUJO.get(mid, ()):

        n = _normalizar(a)

        if n:

            aliases.add(n)

    return aliases





def _emparejar_hoja(nombre_hoja: str) -> str | None:

    """Devuelve entrada.id del catalogo o None."""

    from core.modelos.catalogo_impactos import CATALOGO_MODOS_IMPACTO



    clave = _normalizar(nombre_hoja)

    if not clave:

        return None



    mejor_id: str | None = None

    mejor_score = -1

    for entrada in CATALOGO_MODOS_IMPACTO:

        aliases = _aliases_catalogo(entrada)

        if clave in aliases:

            return entrada.id

        for a in aliases:

            if not a:

                continue

            if a in clave or clave in a:

                score = min(len(a), len(clave))

                if score > mejor_score:

                    mejor_score = score

                    mejor_id = entrada.id

    return mejor_id





def _ext_imagen(data: bytes) -> str:

    if data[:2] == b"\xff\xd8":

        return ".jpg"

    if data[:8].startswith(b"\x89PNG"):

        return ".png"

    if data[:4] == b"GIF8":

        return ".gif"

    return ".bin"





def _data_uri_imagen(data: bytes, ext: str) -> str:

    mime = {

        ".jpg": "image/jpeg",

        ".jpeg": "image/jpeg",

        ".png": "image/png",

        ".gif": "image/gif",

        ".webp": "image/webp",

    }.get(ext.lower(), "application/octet-stream")

    b64 = base64.b64encode(data).decode("ascii")

    return f"data:{mime};base64,{b64}"





def _ancla_desde_img(img) -> tuple[int, int] | None:

    """Devuelve (row, col) 1-based del ancla de una imagen openpyxl."""

    anchor = getattr(img, "anchor", None)

    fr = getattr(anchor, "_from", None)

    if fr is None:

        return None

    return int(fr.row) + 1, int(fr.col) + 1





def _extraer_imagenes_ancladas(ws, carpeta: Path) -> tuple[_ImgAnclada, ...]:

    carpeta.mkdir(parents=True, exist_ok=True)

    out: list[_ImgAnclada] = []

    for i, img in enumerate(list(getattr(ws, "_images", []) or [])):

        try:

            data = img._data()

        except Exception:

            continue

        if not data:

            continue

        ext = _ext_imagen(data)

        destino = carpeta / f"img_{i}{ext}"

        destino.write_bytes(data)

        ancla = _ancla_desde_img(img)

        row, col = ancla if ancla else (1, 1)

        out.append(

            _ImgAnclada(

                row=row,

                col=col,

                path=destino,

                data_uri=_data_uri_imagen(data, ext),

            )

        )

    return tuple(out)





def _limpiar_texto_ecuacion(texto: str) -> str:
    t = texto or ""
    t = t.replace("\u3016", "").replace("\u3017", "")
    t = re.sub(r"[\u200b\u200c\u200d\ufeff\u2061]", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    trans = str.maketrans(
        {
            "\U0001d71f": "\u0394",
            "\U0001d6e5": "\u0394",
            "\U0001d479": "R",
            "\U0001d47a": "S",
            "\U0001d47b": "T",
            "\U0001d46a": "C",
            "\U0001d475": "H",
            "\U0001d473": "F",
            "\U0001d426": "m",
            "\U0001d41a": "a",
            "\U0001d431": "x",
            "\U0001d7ce": "0",
            "\U0001d7d0": "2",
            "\U0001d7d1": "3",
            "\U0001d7d2": "4",
            "\U0001d7d3": "5",
            "\U0001d7d4": "6",
            "\U0001d491": "p",
            "\U0001d490": "o",
            "\U0001d493": "r",
            "\U0001d495": "t",
            "\U0001d484": "c",
            "\U0001d48f": "n",
            "\U0001d460": "s",
            "\U0001d482": "s",
            "\U0001d455": "h",
            "\U0001d456": "i",
            "\U0001d45d": "p",
            "\U0001d45c": "o",
            "\U0001d45f": "r",
            "\U0001d461": "t",
            "\U0001d450": "c",
            "\U0001d45b": "n",
            "\U0001d46f": "H",
            "\U0001d46d": "F",
            "\U0001d494": "s",
            "\U0001d489": "h",
            "\U0001d48a": "i",
            "\u22c5": "\u00b7",
            "\u2212": "-",
        }
    )
    t = t.translate(trans)
    t = re.sub(r"^\u0394R[_]?S\s*=\s*", "\u0394Rs = ", t)
    t = re.sub(r"^\u0394RS\s*=\s*", "\u0394Rs = ", t)
    t = re.sub(r"^\u0394Rs\s*=\s*", "\u0394Rs = ", t)
    t = re.sub(r"\s*=\s*max", " = max", t, flags=re.I)
    t = t.replace("T_port", "Tport").replace("C_conc", "Cconc")
    t = t.replace("(Tport +Cconc)", "(Tport+Cconc)")
    t = re.sub(r"max\{0,\s*", "max{0, ", t)
    return t.strip()


def _texto_desde_omml(fragment: str) -> str:
    """Une texto OMML; prioriza fallback legible de mc:Fallback."""
    fb_block = re.search(r"<mc:Fallback\b[^>]*>(.*?)</mc:Fallback>", fragment, flags=re.S)
    if fb_block:
        texts = re.findall(r"<a:t(?:\s[^>]*)?>(.*?)</a:t>", fb_block.group(1))
        if texts:
            joined = "".join(html_lib.unescape(x) for x in texts)
            limpio = _limpiar_texto_ecuacion(joined)
            if limpio and (
                "max" in limpio.lower() or "\u0394" in limpio or "=" in limpio
            ):
                return limpio

    texts = re.findall(r"<m:t[^>]*>(.*?)</m:t>", fragment)
    if texts:
        return _limpiar_texto_ecuacion("".join(html_lib.unescape(x) for x in texts))

    texts_a = re.findall(r"<a:t(?:\s[^>]*)?>(.*?)</a:t>", fragment)
    if texts_a:
        joined = "".join(html_lib.unescape(x) for x in texts_a if "<" not in x)
        return _limpiar_texto_ecuacion(joined)
    return ""


def _extraer_textos_dibujo_hoja(nombre_hoja: str) -> tuple[_TextoDibujo, ...]:

    """Lee drawing*.xml del xlsx y extrae formas de texto ancladas."""

    if not ARCHIVO_FICHAS.is_file():

        return ()



    import openpyxl



    wb = openpyxl.load_workbook(ARCHIVO_FICHAS, read_only=True, data_only=True)

    try:

        if nombre_hoja not in wb.sheetnames:

            return ()

        sheet_idx = wb.sheetnames.index(nombre_hoja) + 1

    finally:

        wb.close()



    out: list[_TextoDibujo] = []

    try:

        with zipfile.ZipFile(ARCHIVO_FICHAS) as zf:

            rels_name = f"xl/worksheets/_rels/sheet{sheet_idx}.xml.rels"

            if rels_name not in zf.namelist():

                return ()

            rels_xml = zf.read(rels_name).decode("utf-8")

            drawing_targets = re.findall(

                r'Target="([^"]*drawings/drawing[^"]+)"',

                rels_xml,

            )

            for target in drawing_targets:

                path = target.replace("\\", "/")

                if path.startswith("../"):

                    path = "xl/" + path[3:]

                elif not path.startswith("xl/"):

                    path = "xl/worksheets/" + path

                if path not in zf.namelist():

                    continue

                xml = zf.read(path).decode("utf-8")

                for m in re.finditer(

                    r"<xdr:(twoCellAnchor|oneCellAnchor)\b.*?</xdr:\1>",

                    xml,

                    flags=re.S,

                ):

                    chunk = m.group(0)

                    if "<a:t" not in chunk and "<m:t" not in chunk:

                        continue

                    fr = re.search(

                        r"<xdr:from>\s*<xdr:col>(\d+)</xdr:col>.*?"

                        r"<xdr:row>(\d+)</xdr:row>",

                        chunk,

                        flags=re.S,

                    )

                    if not fr:

                        continue

                    col = int(fr.group(1)) + 1

                    row = int(fr.group(2)) + 1

                    texto = _texto_desde_omml(chunk)

                    if texto:

                        out.append(_TextoDibujo(row=row, col=col, texto=texto))

    except Exception:

        return tuple(out)

    return tuple(out)





def _mapa_merges(ws) -> dict[tuple[int, int], tuple[int, int, int, int]]:

    """(row, col) origen -> (min_r, min_c, rowspan, colspan)."""

    out: dict[tuple[int, int], tuple[int, int, int, int]] = {}

    for rango in ws.merged_cells.ranges:

        min_r, min_c, max_r, max_c = rango.min_row, rango.min_col, rango.max_row, rango.max_col

        out[(min_r, min_c)] = (min_r, min_c, max_r - min_r + 1, max_c - min_c + 1)

    return out





def _celdas_cubiertas(ws) -> set[tuple[int, int]]:

    """Celdas que no son el origen de un merge (no deben renderizarse)."""

    cubiertas: set[tuple[int, int]] = set()

    for rango in ws.merged_cells.ranges:

        for r in range(rango.min_row, rango.max_row + 1):

            for c in range(rango.min_col, rango.max_col + 1):

                if r == rango.min_row and c == rango.min_col:

                    continue

                cubiertas.add((r, c))

    return cubiertas





def _texto_celda(valor: object) -> str:

    if valor is None:

        return ""

    return re.sub(r"[\u200b\u200c\u200d\ufeff]", "", str(valor)).strip()





def _fila_tiene_nota_editorial(ws, row: int, max_col: int) -> bool:

    """True si la fila contiene alguna nota editorial (omitir fila entera)."""

    for c in range(1, max_col + 1):

        t = _texto_celda(ws.cell(row, c).value)

        if t and _es_nota_editorial(t):

            return True

    return False





def _limites_hoja(ws) -> tuple[int, int, int, int]:

    """min_row, max_row, min_col, max_col con contenido (tras filtrar notas)."""

    if ws.max_row is None or ws.max_column is None:

        return 1, 0, 1, 0

    max_r = int(ws.max_row)

    max_c = int(ws.max_column)

    filas_utiles: list[int] = []

    cols_utiles: set[int] = set()

    for r in range(1, max_r + 1):

        if _fila_tiene_nota_editorial(ws, r, max_c):

            continue

        tiene = False

        for c in range(1, max_c + 1):

            t = _texto_celda(ws.cell(r, c).value)

            if t and not _es_nota_editorial(t):

                tiene = True

                cols_utiles.add(c)

        if tiene:

            filas_utiles.append(r)

    if not filas_utiles:

        return 1, 0, 1, 0

    for rango in ws.merged_cells.ranges:

        if rango.max_row < filas_utiles[0] or rango.min_row > filas_utiles[-1]:

            continue

        for c in range(rango.min_col, rango.max_col + 1):

            cols_utiles.add(c)

    return filas_utiles[0], filas_utiles[-1], min(cols_utiles), max(cols_utiles)





def _aplicar_tint(rgb_hex: str, tint: float | None) -> str:

    if not tint:

        return rgb_hex

    try:

        r = int(rgb_hex[0:2], 16)

        g = int(rgb_hex[2:4], 16)

        b = int(rgb_hex[4:6], 16)

    except ValueError:

        return rgb_hex

    if tint < 0:

        f = 1.0 + tint

        r, g, b = int(r * f), int(g * f), int(b * f)

    else:

        r = int(r + (255 - r) * tint)

        g = int(g + (255 - g) * tint)

        b = int(b + (255 - b) * tint)

    return f"{max(0, min(255, r)):02X}{max(0, min(255, g)):02X}{max(0, min(255, b)):02X}"





def _color_openpyxl(color) -> str | None:

    """Devuelve #RRGGBB o None."""

    if color is None or getattr(color, "type", None) is None:

        return None

    ctype = color.type

    if ctype == "rgb" and color.rgb:

        rgb = str(color.rgb)

        if rgb.startswith("0x"):

            rgb = rgb[2:]

        if len(rgb) == 8:

            rgb = rgb[2:]

        if len(rgb) == 6:

            return f"#{rgb.upper()}"

    if ctype == "theme" and color.theme is not None:

        base = _THEME_RGB.get(int(color.theme))

        if not base:

            return None

        tint = getattr(color, "tint", None) or 0.0

        return f"#{_aplicar_tint(base, float(tint))}"

    return None





def _estilo_celda(cell) -> str:

    """CSS inline a partir de fill/font/alignment de openpyxl."""

    parts: list[str] = []

    fill = cell.fill

    if fill is not None and fill.fill_type == "solid":

        bg = _color_openpyxl(fill.fgColor) or _color_openpyxl(fill.start_color)

        if bg and bg.upper() not in {"#FFFFFF", "#00000000"}:

            parts.append(f"background-color:{bg}")



    font = cell.font

    if font is not None:

        if font.bold:

            parts.append("font-weight:700")

        if font.italic:

            parts.append("font-style:italic")

        fc = _color_openpyxl(font.color)

        if fc:

            parts.append(f"color:{fc}")

        if font.size:

            parts.append(f"font-size:{font.size}pt")



    align = cell.alignment

    if align is not None:

        h = (align.horizontal or "").lower()

        v = (align.vertical or "").lower()

        if h in {"center", "centre"}:

            parts.append("text-align:center")

        elif h == "right":

            parts.append("text-align:right")

        elif h == "justify":

            parts.append("text-align:justify")

        elif h == "left":

            parts.append("text-align:left")

        if v in {"center", "centre"}:

            parts.append("vertical-align:middle")

        elif v == "top":

            parts.append("vertical-align:top")

        elif v == "bottom":

            parts.append("vertical-align:bottom")

        if align.wrap_text:

            parts.append("white-space:pre-wrap")

        rot = getattr(align, "textRotation", None) or 0

        if rot in (90, 255):

            parts.append("writing-mode:vertical-rl")

            parts.append("transform:rotate(180deg)")

    return ";".join(parts)





def _html_desde_hoja(

    ws,

    *,

    imagenes: tuple[_ImgAnclada, ...] = (),

    textos_dibujo: tuple[_TextoDibujo, ...] = (),

) -> tuple[str, tuple[Path, ...]]:

    """Tabla HTML con estilos Excel. Devuelve (html, imagenes_esquema)."""

    min_r, max_r, min_c, max_c = _limites_hoja(ws)

    if max_r < min_r or max_c < min_c:

        return "", tuple(img.path for img in imagenes)



    merges = _mapa_merges(ws)

    cubiertas = _celdas_cubiertas(ws)

    max_col_ws = int(ws.max_column or 0)

    filas_omitir = {

        r

        for r in range(1, int(ws.max_row or 0) + 1)

        if _fila_tiene_nota_editorial(ws, r, max_col_ws)

    }



    imgs_tabla: dict[tuple[int, int], list[_ImgAnclada]] = {}

    imgs_esquema: list[Path] = []

    for img in imagenes:

        # Esquema a la derecha del bloque tabular, o fuera de filas de contenido.

        if img.col > max_c or img.row < min_r:

            imgs_esquema.append(img.path)

        elif min_r <= img.row <= max_r + 1 and min_c <= img.col <= max_c:

            imgs_tabla.setdefault((img.row, img.col), []).append(img)

        else:

            imgs_esquema.append(img.path)



    textos_por_celda: dict[tuple[int, int], list[str]] = {}

    for td in textos_dibujo:

        if td.row < min_r or td.row > max_r + 2:

            continue

        if td.col < min_c or td.col > max_c + 1:

            continue

        textos_por_celda.setdefault((td.row, td.col), []).append(td.texto)



    filas_html: list[str] = []

    for r in range(min_r, max_r + 1):

        if r in filas_omitir:

            continue

        celdas: list[str] = []

        for c in range(min_c, max_c + 1):

            if (r, c) in cubiertas:

                continue

            cell = ws.cell(r, c)

            texto = _texto_celda(cell.value)

            if _es_nota_editorial(texto):

                texto = ""

            attrs = ""

            info = merges.get((r, c))

            rowspan_eff = 1

            if info is not None:

                _mr, _mc, rowspan, colspan = info

                fin = _mr + rowspan - 1

                omitidas = sum(1 for rr in range(_mr, fin + 1) if rr in filas_omitir)

                rowspan_eff = max(1, rowspan - omitidas)

                if rowspan_eff > 1:

                    attrs += f' rowspan="{rowspan_eff}"'

                if colspan > 1:

                    attrs += f' colspan="{colspan}"'



            style = _estilo_celda(cell)

            if rowspan_eff > 1 and c == min_c and texto:

                extra = (

                    "font-weight:700;text-align:center;"

                    "vertical-align:middle;background-color:#f3f5f7"

                )

                style = f"{style};{extra}" if style else extra



            extras_html: list[str] = []

            coords_merge = {(r, c)}

            if info is not None:

                _mr, _mc, rowspan, colspan = info

                for rr in range(_mr, _mr + rowspan):

                    for cc in range(_mc, _mc + colspan):

                        coords_merge.add((rr, cc))

            for coord in coords_merge:

                for t in textos_por_celda.pop(coord, []):

                    extras_html.append(

                        f'<div class="pde-ficha-eq">{html_lib.escape(t)}</div>'

                    )

                for im in imgs_tabla.pop(coord, []):

                    extras_html.append(

                        f'<div class="pde-ficha-eq-img">'

                        f'<img src="{im.data_uri}" alt="Ecuacion"/></div>'

                    )



            esc = html_lib.escape(texto).replace("\n", "<br/>")

            inner = esc + ("".join(extras_html) if extras_html else "")

            style_attr = f' style="{style}"' if style else ""

            celdas.append(f"<td{attrs}{style_attr}>{inner}</td>")

        if celdas:

            filas_html.append(f"<tr>{''.join(celdas)}</tr>")



    pendientes: list[str] = []

    for textos in textos_por_celda.values():

        for t in textos:

            pendientes.append(html_lib.escape(t))

    for ims in imgs_tabla.values():

        for im in ims:

            pendientes.append(

                f'<img src="{im.data_uri}" alt="Ecuacion" '

                f'style="max-width:100%;height:auto;"/>'

            )

    if pendientes and filas_html:

        eqs = "".join(f'<div class="pde-ficha-eq">{p}</div>' for p in pendientes)

        span = max_c - min_c + 1

        filas_html.append(

            f'<tr><td colspan="{span}" '

            f'style="text-align:center;padding:12px 10px;">{eqs}</td></tr>'

        )



    if not filas_html:

        return "", tuple(imgs_esquema)



    html = f"""

<div class="pde-ficha-excel">

<style>

.pde-ficha-excel table {{

  border-collapse: collapse;

  width: 100%;

  font-size: 0.88rem;

  table-layout: fixed;

  background: #fff;

}}

.pde-ficha-excel th,

.pde-ficha-excel td {{

  border: 1px solid #1a1a1a;

  padding: 8px 10px;

  text-align: left;

  vertical-align: middle;

  word-wrap: break-word;

}}

.pde-ficha-excel .pde-ficha-eq {{

  display: block;

  text-align: center;

  font-weight: 700;

  font-style: italic;

  margin-top: 10px;

  font-size: 0.95rem;

  letter-spacing: 0.01em;

}}

.pde-ficha-excel .pde-ficha-eq-img {{

  text-align: center;

  margin-top: 8px;

}}

.pde-ficha-excel .pde-ficha-eq-img img {{

  max-width: 100%;

  height: auto;

}}

</style>

<table>

  <tbody>

    {''.join(filas_html)}

  </tbody>

</table>

</div>

""".strip()

    return html, tuple(imgs_esquema)





def _tabla_desde_hoja(ws) -> pd.DataFrame:

    """DataFrame plano (fallback / depuracion); omite notas editoriales."""

    min_r, max_r, min_c, max_c = _limites_hoja(ws)

    if max_r < min_r:

        return pd.DataFrame()

    cubiertas = _celdas_cubiertas(ws)

    max_col_ws = int(ws.max_column or 0)

    filas_omitir = {

        r

        for r in range(1, int(ws.max_row or 0) + 1)

        if _fila_tiene_nota_editorial(ws, r, max_col_ws)

    }

    filas: list[list[str]] = []

    n_cols = max_c - min_c + 1

    for r in range(min_r, max_r + 1):

        if r in filas_omitir:

            continue

        vals = [""] * n_cols

        for c in range(min_c, max_c + 1):

            if (r, c) in cubiertas:

                continue

            texto = _texto_celda(ws.cell(r, c).value)

            if _es_nota_editorial(texto):

                texto = ""

            vals[c - min_c] = texto

        if any(vals):

            filas.append(vals)

    if not filas:

        return pd.DataFrame()

    columnas = [f"Columna {j + 1}" for j in range(n_cols)]

    return pd.DataFrame(filas, columns=columnas)





@lru_cache(maxsize=1)

def _cargar_fichas_excel() -> dict[str, FichaExcelModelo]:

    """Mapa catalogo_id -> ficha. Si no hay match, clave hoja:<nombre>."""

    if not ARCHIVO_FICHAS.is_file():

        return {}



    import openpyxl



    wb = openpyxl.load_workbook(ARCHIVO_FICHAS, data_only=False)

    resultado: dict[str, FichaExcelModelo] = {}

    for nombre in wb.sheetnames:

        ws = wb[nombre]

        catalogo_id = _emparejar_hoja(nombre)

        media = CARPETA_MEDIA / _slug(nombre)

        imagenes_anc = _extraer_imagenes_ancladas(ws, media)

        textos = _extraer_textos_dibujo_hoja(nombre)

        html, imgs_esquema = _html_desde_hoja(

            ws, imagenes=imagenes_anc, textos_dibujo=textos

        )

        if not imgs_esquema and imagenes_anc and "data:image" not in html:

            imgs_esquema = tuple(img.path for img in imagenes_anc)

        key = catalogo_id or f"hoja:{nombre}"

        resultado[key] = FichaExcelModelo(

            hoja=nombre,

            tabla=_tabla_desde_hoja(ws),

            html=html,

            imagenes=imgs_esquema,

            catalogo_id=catalogo_id,

        )

    wb.close()

    return resultado





def invalidar_cache_fichas() -> None:

    _cargar_fichas_excel.cache_clear()





def ficha_excel_por_catalogo_id(catalogo_id: str) -> FichaExcelModelo | None:

    if not (catalogo_id or "").strip():

        return None

    return _cargar_fichas_excel().get(catalogo_id.strip())





def ficha_excel_por_entrada(entrada) -> FichaExcelModelo | None:

    """Resuelve ficha Excel para una entrada del catalogo."""

    if entrada is None:

        return None

    directa = ficha_excel_por_catalogo_id(entrada.id)

    if directa is not None:

        return directa

    for ficha in _cargar_fichas_excel().values():

        if ficha.catalogo_id == entrada.id:

            return ficha

        if _emparejar_hoja(ficha.hoja) == entrada.id:

            return ficha

    return None





def imagenes_ficha_por_entrada(entrada) -> tuple[Path, ...]:

    ficha = ficha_excel_por_entrada(entrada)

    if ficha is None:

        return ()

    return ficha.imagenes

