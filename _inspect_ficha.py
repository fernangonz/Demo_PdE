# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook("Fichas/Ficha.xlsx", data_only=False)
print("sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print("===", name, "max", ws.max_row, ws.max_column)
    print("merges:", list(ws.merged_cells.ranges))
    imgs = list(getattr(ws, "_images", []) or [])
    print("images:", len(imgs))
    for i, img in enumerate(imgs):
        anchor = getattr(img, "anchor", None)
        print("  img", i, type(img).__name__, anchor)
        fr = getattr(anchor, "_from", None)
        if fr is not None:
            print("    from", fr.col, fr.row)
            to = getattr(anchor, "_to", None)
            if to is not None:
                print("    to", to.col, to.row)
        print("    size", getattr(img, "width", None), getattr(img, "height", None))
    for r in range(1, min(ws.max_row or 0, 30) + 1):
        for c in range(1, min(ws.max_column or 0, 8) + 1):
            cell = ws.cell(r, c)
            fill = cell.fill
            font = cell.font
            align = cell.alignment
            fill_info = None
            if fill and fill.fgColor and fill.fill_type:
                if fill.fgColor.type == "rgb":
                    rgb = fill.fgColor.rgb
                else:
                    rgb = "theme:%s" % fill.fgColor.theme
                fill_info = (fill.fill_type, rgb)
            font_info = None
            if font:
                fc = None
                if font.color:
                    if font.color.type == "rgb":
                        fc = font.color.rgb
                    else:
                        fc = "theme:%s" % font.color.theme
                font_info = (font.bold, fc, font.size)
            val = repr(cell.value)[:100] if cell.value is not None else None
            if val is None and not fill_info:
                continue
            al = None
            if align:
                al = (align.horizontal, align.vertical, align.wrap_text)
            print("[%s,%s] val=%s fill=%s font=%s align=%s" % (r, c, val, fill_info, font_info, al))
